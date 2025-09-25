from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import io
import logging
from typing import Optional
import re

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="ERP to Core Tax Converter", version="1.0.0")

class CoreTaxConverter:
    def __init__(self):
        self.ppn_rate = 0.12  # 12% PPN rate
        
    def clean_numeric_value(self, value):
        """Clean and convert string numbers to float, ensuring no NaN values"""
        # Handle None, NaN, empty string, or missing values
        if value is None or pd.isna(value) or value == '' or value == 'NaN':
            return 0.0
            
        # Handle numpy NaN or Python float NaN
        if isinstance(value, float) and (value != value or str(value).lower() == 'nan'):  # NaN != NaN is True
            return 0.0
            
        if isinstance(value, str):
            # Check for string 'nan' or 'NaN'
            if value.lower().strip() == 'nan':
                return 0.0
            # Remove any non-numeric characters except decimal points
            cleaned = re.sub(r'[^\d.-]', '', str(value))
            if not cleaned or cleaned == '-' or cleaned == '.':
                return 0.0
            try:
                result = float(cleaned)
                # Double-check result is not NaN
                return result if not (result != result) else 0.0
            except (ValueError, OverflowError):
                return 0.0
                
        try:
            result = float(value) if value else 0.0
            # Ensure result is not NaN or infinity
            if result != result or result == float('inf') or result == float('-inf'):
                return 0.0
            return result
        except (ValueError, TypeError, OverflowError):
            return 0.0
    
    def calculate_dpp_and_ppn(self, price_after_tax):
        """Calculate DPP and PPN from price after tax, ensuring no NaN values"""
        try:
            # Clean the input value first
            price_after_tax = self.clean_numeric_value(price_after_tax)
            
            if price_after_tax <= 0:
                return 0.0, 0.0
            
            # DPP = Price After Tax / (1 + PPN Rate)
            denominator = 1 + self.ppn_rate
            if denominator == 0:  # Safety check (should never happen)
                return 0.0, 0.0
                
            dpp = price_after_tax / denominator
            ppn = dpp * self.ppn_rate
            
            # Ensure results are valid numbers
            dpp = dpp if (dpp == dpp and dpp != float('inf') and dpp != float('-inf')) else 0.0
            ppn = ppn if (ppn == ppn and ppn != float('inf') and ppn != float('-inf')) else 0.0
            
            return round(dpp, 2), round(ppn, 2)
            
        except (ValueError, TypeError, ZeroDivisionError, OverflowError) as e:
            logger.error(f"Error in DPP/PPN calculation: {e}")
            return 0.0, 0.0
    
    def process_sales_data(self, sales_df):
        """Process sales data and convert to Core Tax format with NaN prevention"""
        logger.info(f"Processing {len(sales_df)} sales records")
        
        # Clean column names
        sales_df.columns = sales_df.columns.str.strip()
        
        processed_data = []
        
        for idx, row in sales_df.iterrows():
            try:
                # Extract and clean basic info
                customer_code = str(row.get('CustomerCode', '')).strip()
                customer_name = str(row.get('CustomerName', '')).strip()
                invoice_no = str(row.get('InvoiceNo', '')).strip()
                invoice_date = row.get('InvoiceDate', '')
                item_code = str(row.get('ItemCode', '')).strip()
                item_name = str(row.get('ItemName', '')).strip()
                
                # Handle quantity and pricing with NaN prevention
                qty = self.clean_numeric_value(row.get('Qty', 0))
                price_after_tax = self.clean_numeric_value(row.get('PriceAfterTax', 0))
                invoice_amount = self.clean_numeric_value(row.get('InvoiceAmount', 0))
                
                # Ensure minimum values to prevent division by zero
                qty = max(qty, 1)  # Minimum quantity of 1
                
                # Use invoice amount if available and valid, otherwise use price after tax
                total_amount = invoice_amount if invoice_amount > 0 else price_after_tax
                total_amount = max(total_amount, 0)  # Ensure non-negative
                
                # Calculate unit price with zero-division protection
                unit_price = (total_amount / qty) if qty > 0 else 0
                
                # Calculate DPP and PPN
                dpp_total, ppn_total = self.calculate_dpp_and_ppn(total_amount)
                dpp_unit = (dpp_total / qty) if qty > 0 else 0
                
                # Format invoice date
                formatted_date = self.format_date(invoice_date)
                
                # Create record with all NaN-safe values
                record = {
                    'baris': idx + 1,
                    'barang_jasa': 'A',  # Default to 'A' for goods
                    'kode_barang_jasa': item_code[:20] if item_code else '310000',  # Default code
                    'nama_barang_jasa': item_name[:255] if item_name else 'Barang/Jasa',
                    'nama_satuan_ukur': 'UM.0003',  # Default unit
                    'harga_satuan': self.safe_round(dpp_unit, 2),
                    'jumlah_barang_jasa': int(qty) if qty > 0 else 1,
                    'total_diskon': 0.0,
                    'dpp': self.safe_round(dpp_total, 2),
                    'dpp_nilai_lain': self.safe_round(dpp_total, 2),
                    'tarif_ppn': 12,
                    'ppn': self.safe_round(ppn_total, 2),
                    'tarif_ppnbm': 0,
                    'ppnbm': 0.0,
                    'customer_code': customer_code,
                    'customer_name': customer_name,
                    'invoice_no': invoice_no,
                    'invoice_date': formatted_date,
                    'total_amount': self.safe_round(total_amount, 2)
                }
                
                # Final validation - ensure no field contains NaN
                validated_record = self.validate_record(record)
                processed_data.append(validated_record)
                
            except Exception as e:
                logger.error(f"Error processing row {idx}: {str(e)}")
                # Create a minimal valid record to prevent data loss
                fallback_record = self.create_fallback_record(idx + 1)
                processed_data.append(fallback_record)
                continue
        
        logger.info(f"Successfully processed {len(processed_data)} records")
        return processed_data
    
    def safe_round(self, value, decimals=2):
        """Safely round a number, handling NaN and infinity"""
        try:
            if value is None or value != value or value == float('inf') or value == float('-inf'):
                return 0.0
            return round(float(value), decimals)
        except (ValueError, TypeError, OverflowError):
            return 0.0
    
    def validate_record(self, record):
        """Validate and clean a record to ensure no NaN values"""
        validated = {}
        for key, value in record.items():
            if isinstance(value, (int, float)):
                # Check for NaN, infinity, or invalid numbers
                if value != value or value == float('inf') or value == float('-inf'):
                    validated[key] = 0.0 if key in ['harga_satuan', 'dpp', 'dpp_nilai_lain', 'ppn', 'ppnbm', 'total_diskon', 'total_amount'] else 0
                else:
                    validated[key] = value
            elif isinstance(value, str):
                # Ensure no 'NaN' strings
                validated[key] = value if value.lower() != 'nan' else ''
            else:
                validated[key] = value
        return validated
    
    def create_fallback_record(self, row_number):
        """Create a minimal valid record as fallback"""
        return {
            'baris': row_number,
            'barang_jasa': 'A',
            'kode_barang_jasa': '310000',
            'nama_barang_jasa': 'Data Error - Manual Review Required',
            'nama_satuan_ukur': 'UM.0003',
            'harga_satuan': 0.0,
            'jumlah_barang_jasa': 1,
            'total_diskon': 0.0,
            'dpp': 0.0,
            'dpp_nilai_lain': 0.0,
            'tarif_ppn': 12,
            'ppn': 0.0,
            'tarif_ppnbm': 0,
            'ppnbm': 0.0,
            'customer_code': '',
            'customer_name': '',
            'invoice_no': '',
            'invoice_date': datetime.now().strftime('%Y-%m-%d'),
            'total_amount': 0.0
        }
    
    def format_date(self, date_value):
        """Format date to YYYY-MM-DD"""
        if pd.isna(date_value):
            return datetime.now().strftime('%Y-%m-%d')
        
        if isinstance(date_value, str):
            # Try to parse various date formats
            date_formats = ['%d.%m.%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(date_value, fmt)
                    return parsed_date.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        
        return datetime.now().strftime('%Y-%m-%d')
    
    def create_core_tax_excel(self, processed_data, company_npwp="0012328415631000"):
        """Create Excel file in Core Tax format"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Faktur sheet
        faktur_sheet = wb.create_sheet("Faktur")
        self.create_faktur_sheet(faktur_sheet, company_npwp)
        
        # Create DetailFaktur sheet
        detail_sheet = wb.create_sheet("DetailFaktur")
        self.create_detail_faktur_sheet(detail_sheet, processed_data)
        
        # Create REF sheet
        ref_sheet = wb.create_sheet("REF")
        self.create_ref_sheet(ref_sheet)
        
        # Create Keterangan sheet
        keterangan_sheet = wb.create_sheet("Keterangan")
        self.create_keterangan_sheet(keterangan_sheet)
        
        return wb
    
    def create_faktur_sheet(self, sheet, npwp):
        """Create the Faktur header sheet"""
        sheet['A1'] = 'NPWP Penjual'
        sheet['C1'] = npwp
        
        sheet['A3'] = 'Baris'
        sheet['C3'] = 'Jenis Faktur'
        sheet['E3'] = 'Keterangan Tambahan'
        
        # Add some sample rows
        for i in range(4, 9):
            sheet[f'A{i}'] = i - 3
            sheet[f'C{i}'] = 'Normal'
    
    def create_detail_faktur_sheet(self, sheet, processed_data):
        """Create the DetailFaktur sheet with transaction data, ensuring no NaN values"""
        # Headers
        headers = [
            'Baris', 'Barang.Jasa', 'Kode Barang Jasa', 'Nama Barang.Jasa', 
            'Nama Satuan Ukur', 'Harga Satuan', 'Jumlah Barang Jasa', 'Total Diskon',
            'DPP', 'DPP Nilai Lain', 'Tarif PPN', 'PPN', 'Tarif PPnBM', 'PPnBM'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Write data with NaN prevention
        for row_idx, record in enumerate(processed_data, 2):
            try:
                # Write each cell with validation
                self.safe_write_cell(sheet, row_idx, 1, record['baris'])
                self.safe_write_cell(sheet, row_idx, 2, record['barang_jasa'])
                self.safe_write_cell(sheet, row_idx, 3, record['kode_barang_jasa'])
                self.safe_write_cell(sheet, row_idx, 4, record['nama_barang_jasa'])
                self.safe_write_cell(sheet, row_idx, 5, record['nama_satuan_ukur'])
                self.safe_write_cell(sheet, row_idx, 6, record['harga_satuan'])
                self.safe_write_cell(sheet, row_idx, 7, record['jumlah_barang_jasa'])
                self.safe_write_cell(sheet, row_idx, 8, record['total_diskon'])
                self.safe_write_cell(sheet, row_idx, 9, record['dpp'])
                self.safe_write_cell(sheet, row_idx, 10, record['dpp_nilai_lain'])
                self.safe_write_cell(sheet, row_idx, 11, record['tarif_ppn'])
                self.safe_write_cell(sheet, row_idx, 12, record['ppn'])
                self.safe_write_cell(sheet, row_idx, 13, record['tarif_ppnbm'])
                self.safe_write_cell(sheet, row_idx, 14, record['ppnbm'])
                
            except Exception as e:
                logger.error(f"Error writing row {row_idx}: {str(e)}")
                # Write fallback values for this row
                for col in range(1, 15):
                    fallback_value = 0 if col in [6, 8, 9, 10, 12, 14] else (1 if col == 7 else ('A' if col == 2 else ''))
                    self.safe_write_cell(sheet, row_idx, col, fallback_value)
    
    def safe_write_cell(self, sheet, row, col, value):
        """Safely write a value to a cell, preventing NaN values"""
        try:
            # Handle different value types
            if isinstance(value, (int, float)):
                # Check for NaN, infinity, or invalid numbers
                if value != value or value == float('inf') or value == float('-inf'):
                    safe_value = 0
                else:
                    safe_value = value
            elif isinstance(value, str):
                # Check for 'NaN' strings
                safe_value = value if value.lower() != 'nan' else ''
            elif value is None:
                safe_value = 0 if col in [6, 8, 9, 10, 12, 14] else ''  # Numeric columns get 0, text get empty
            else:
                safe_value = value
            
            sheet.cell(row=row, column=col, value=safe_value)
            
        except Exception as e:
            logger.error(f"Error writing cell ({row}, {col}): {str(e)}")
            # Write a safe fallback value
            fallback_value = 0 if col in [1, 6, 7, 8, 9, 10, 11, 12, 13, 14] else ''
            sheet.cell(row=row, column=col, value=fallback_value)
    
    def create_ref_sheet(self, sheet):
        """Create reference sheet"""
        sheet['A1'] = 'Kode'
        sheet['B1'] = 'Keterangan'
        sheet['A2'] = 'Barang/Jasa'
    
    def create_keterangan_sheet(self, sheet):
        """Create explanation sheet"""
        headers = ['Kolom', 'Mandatory', 'Validasi DJP', 'Keterangan']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        sheet['A2'] = 'Faktur'

# Initialize converter
converter = CoreTaxConverter()

@app.get("/", response_class=HTMLResponse)
async def read_root():
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>ERP to Core Tax Converter</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                color: #333;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
            }
            
            .container {
                background: white;
                padding: 2rem;
                border-radius: 15px;
                box-shadow: 0 20px 40px rgba(0,0,0,0.1);
                max-width: 600px;
                width: 90%;
                text-align: center;
            }
            
            .header {
                margin-bottom: 2rem;
            }
            
            .header h1 {
                color: #2c3e50;
                margin-bottom: 0.5rem;
                font-size: 2rem;
            }
            
            .header p {
                color: #7f8c8d;
                font-size: 1.1rem;
            }
            
            .upload-area {
                border: 3px dashed #bdc3c7;
                border-radius: 10px;
                padding: 3rem 1rem;
                margin: 2rem 0;
                transition: all 0.3s ease;
                cursor: pointer;
                position: relative;
                overflow: hidden;
            }
            
            .upload-area:hover {
                border-color: #3498db;
                background-color: #f8f9fa;
                transform: translateY(-2px);
            }
            
            .upload-area.dragover {
                border-color: #2ecc71;
                background-color: #d5f4e6;
            }
            
            .upload-icon {
                font-size: 3rem;
                color: #bdc3c7;
                margin-bottom: 1rem;
            }
            
            .upload-text {
                font-size: 1.2rem;
                margin-bottom: 1rem;
                color: #2c3e50;
            }
            
            .upload-subtext {
                color: #7f8c8d;
                font-size: 0.9rem;
            }
            
            .file-input {
                display: none;
            }
            
            .btn {
                background: linear-gradient(135deg, #3498db, #2980b9);
                color: white;
                padding: 12px 30px;
                border: none;
                border-radius: 25px;
                cursor: pointer;
                font-size: 1rem;
                font-weight: 600;
                transition: all 0.3s ease;
                margin: 0.5rem;
                text-decoration: none;
                display: inline-block;
            }
            
            .btn:hover {
                transform: translateY(-2px);
                box-shadow: 0 5px 15px rgba(52, 152, 219, 0.4);
            }
            
            .btn:disabled {
                background: #bdc3c7;
                cursor: not-allowed;
                transform: none;
                box-shadow: none;
            }
            
            .progress {
                margin: 2rem 0;
                display: none;
            }
            
            .progress-bar {
                background-color: #ecf0f1;
                border-radius: 10px;
                overflow: hidden;
                height: 20px;
            }
            
            .progress-fill {
                background: linear-gradient(90deg, #2ecc71, #27ae60);
                height: 100%;
                width: 0%;
                transition: width 0.3s ease;
                border-radius: 10px;
            }
            
            .progress-text {
                margin-top: 0.5rem;
                color: #2c3e50;
                font-weight: 600;
            }
            
            .file-info {
                background: #f8f9fa;
                padding: 1rem;
                border-radius: 8px;
                margin: 1rem 0;
                display: none;
            }
            
            .file-info h3 {
                color: #2c3e50;
                margin-bottom: 0.5rem;
            }
            
            .file-info p {
                color: #7f8c8d;
                margin: 0.25rem 0;
            }
            
            .download-section {
                background: #e8f5e8;
                padding: 1.5rem;
                border-radius: 10px;
                margin: 1rem 0;
                display: none;
            }
            
            .download-section h3 {
                color: #27ae60;
                margin-bottom: 1rem;
            }
            
            .btn-success {
                background: linear-gradient(135deg, #2ecc71, #27ae60);
            }
            
            .btn-success:hover {
                box-shadow: 0 5px 15px rgba(46, 204, 113, 0.4);
            }
            
            .error {
                background: #fee;
                color: #c0392b;
                padding: 1rem;
                border-radius: 8px;
                margin: 1rem 0;
                display: none;
                border-left: 4px solid #e74c3c;
            }
            
            .success {
                background: #efe;
                color: #27ae60;
                padding: 1rem;
                border-radius: 8px;
                margin: 1rem 0;
                display: none;
                border-left: 4px solid #2ecc71;
            }
            
            .features {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 1rem;
                margin: 2rem 0;
            }
            
            .feature {
                background: #f8f9fa;
                padding: 1.5rem;
                border-radius: 10px;
                text-align: left;
            }
            
            .feature h4 {
                color: #2c3e50;
                margin-bottom: 0.5rem;
                font-size: 1.1rem;
            }
            
            .feature p {
                color: #7f8c8d;
                font-size: 0.9rem;
            }
            
            .footer {
                margin-top: 2rem;
                padding-top: 1rem;
                border-top: 1px solid #ecf0f1;
                color: #7f8c8d;
                font-size: 0.9rem;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🏢 ERP to Core Tax Converter</h1>
                <p>Convert your ERP sales data to Core Tax import format</p>
            </div>
            
            <div class="features">
                <div class="feature">
                    <h4>📊 Smart Mapping</h4>
                    <p>Automatically maps your sales data to Core Tax format</p>
                </div>
                <div class="feature">
                    <h4>💰 Tax Calculation</h4>
                    <p>Calculates DPP and PPN automatically based on your data</p>
                </div>
                <div class="feature">
                    <h4>⚡ Fast Processing</h4>
                    <p>Process hundreds of transactions in seconds</p>
                </div>
            </div>
            
            <div class="upload-area" onclick="document.getElementById('fileInput').click()">
                <div class="upload-icon">📁</div>
                <div class="upload-text">Click to upload or drag & drop</div>
                <div class="upload-subtext">Upload your Sales.xlsx file</div>
                <input type="file" id="fileInput" class="file-input" accept=".xlsx,.xls" />
            </div>
            
            <div class="file-info" id="fileInfo">
                <h3>File Information</h3>
                <p id="fileName"></p>
                <p id="fileSize"></p>
                <p id="recordCount"></p>
            </div>
            
            <div class="progress" id="progress">
                <div class="progress-bar">
                    <div class="progress-fill" id="progressFill"></div>
                </div>
                <div class="progress-text" id="progressText">Processing...</div>
            </div>
            
            <div class="error" id="error"></div>
            <div class="success" id="success"></div>
            
            <div class="download-section" id="downloadSection">
                <h3>✅ Conversion Complete!</h3>
                <p>Your file has been converted to Core Tax format successfully.</p>
                <a href="#" id="downloadLink" class="btn btn-success">📥 Download Core Tax File</a>
            </div>
            
            <button class="btn" id="convertBtn" style="display: none;" onclick="convertFile()">
                🔄 Convert to Core Tax Format
            </button>
            
            <div class="footer">
                <p>Supports Excel files (.xlsx, .xls) with sales transaction data</p>
                <p>Generates Core Tax compatible format for easy import</p>
            </div>
        </div>
        
        <script>
            let uploadedFile = null;
            
            // File upload handling
            const fileInput = document.getElementById('fileInput');
            const uploadArea = document.querySelector('.upload-area');
            const fileInfo = document.getElementById('fileInfo');
            const convertBtn = document.getElementById('convertBtn');
            const progress = document.getElementById('progress');
            const progressFill = document.getElementById('progressFill');
            const progressText = document.getElementById('progressText');
            const error = document.getElementById('error');
            const success = document.getElementById('success');
            const downloadSection = document.getElementById('downloadSection');
            const downloadLink = document.getElementById('downloadLink');
            
            // Drag and drop handlers
            uploadArea.addEventListener('dragover', (e) => {
                e.preventDefault();
                uploadArea.classList.add('dragover');
            });
            
            uploadArea.addEventListener('dragleave', () => {
                uploadArea.classList.remove('dragover');
            });
            
            uploadArea.addEventListener('drop', (e) => {
                e.preventDefault();
                uploadArea.classList.remove('dragover');
                const files = e.dataTransfer.files;
                if (files.length > 0) {
                    handleFile(files[0]);
                }
            });
            
            fileInput.addEventListener('change', (e) => {
                if (e.target.files.length > 0) {
                    handleFile(e.target.files[0]);
                }
            });
            
            function handleFile(file) {
                if (!file.name.match(/\.(xlsx?|xls)$/i)) {
                    showError('Please select an Excel file (.xlsx or .xls)');
                    return;
                }
                
                uploadedFile = file;
                
                // Show file info
                document.getElementById('fileName').textContent = `File: ${file.name}`;
                document.getElementById('fileSize').textContent = `Size: ${(file.size / 1024 / 1024).toFixed(2)} MB`;
                
                fileInfo.style.display = 'block';
                convertBtn.style.display = 'inline-block';
                hideMessages();
            }
            
            async function convertFile() {
                if (!uploadedFile) {
                    showError('Please select a file first');
                    return;
                }
                
                const formData = new FormData();
                formData.append('file', uploadedFile);
                
                // Show progress
                showProgress();
                convertBtn.disabled = true;
                
                try {
                    const response = await fetch('/convert/', {
                        method: 'POST',
                        body: formData
                    });
                    
                    if (!response.ok) {
                        const errorData = await response.json();
                        throw new Error(errorData.detail || 'Conversion failed');
                    }
                    
                    // Update progress
                    updateProgress(100, 'Conversion complete!');
                    
                    // Create download link
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    downloadLink.href = url;
                    downloadLink.download = 'CoreTax_Import_' + new Date().toISOString().slice(0, 19).replace(/:/g, '-') + '.xlsx';
                    
                    // Show success
                    hideProgress();
                    downloadSection.style.display = 'block';
                    showSuccess('File converted successfully! Click the download button to get your Core Tax format file.');
                    
                } catch (err) {
                    hideProgress();
                    showError('Error: ' + err.message);
                } finally {
                    convertBtn.disabled = false;
                }
            }
            
            function showProgress() {
                progress.style.display = 'block';
                updateProgress(0, 'Starting conversion...');
                
                // Simulate progress
                let progressValue = 0;
                const interval = setInterval(() => {
                    progressValue += Math.random() * 20;
                    if (progressValue < 90) {
                        updateProgress(progressValue, 'Processing data...');
                    } else {
                        clearInterval(interval);
                        updateProgress(90, 'Finalizing...');
                    }
                }, 200);
            }
            
            function updateProgress(value, text) {
                progressFill.style.width = value + '%';
                progressText.textContent = text;
            }
            
            function hideProgress() {
                progress.style.display = 'none';
            }
            
            function showError(message) {
                error.textContent = message;
                error.style.display = 'block';
                success.style.display = 'none';
            }
            
            function showSuccess(message) {
                success.textContent = message;
                success.style.display = 'block';
                error.style.display = 'none';
            }
            
            function hideMessages() {
                error.style.display = 'none';
                success.style.display = 'none';
            }
        </script>
    </body>
    </html>
    """
    return html_content

@app.post("/convert/")
async def convert_file(file: UploadFile = File(...)):
    """Convert uploaded Excel file to Core Tax format"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        # Read the uploaded file
        contents = await file.read()
        
        # Load into pandas DataFrame
        df = pd.read_excel(io.BytesIO(contents))
        logger.info(f"Loaded {len(df)} records from {file.filename}")
        
        # Process the data
        processed_data = converter.process_sales_data(df)
        
        if not processed_data:
            raise HTTPException(status_code=400, detail="No valid data found in the uploaded file")
        
        # Create Core Tax Excel file
        workbook = converter.create_core_tax_excel(processed_data)
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Return as streaming response
        filename = f"CoreTax_Import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error converting file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
